#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 读取JSON数据
with open('pxb7_products_50.json', 'r', encoding='utf-8') as f:
    products = json.load(f)

# 创建工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = "王者荣耀世界账号交易"

# 设置表头
headers = ["序号", "商品名称", "平台", "价格", "发布时间", "商品链接"]
ws.append(headers)

# 设置表头样式
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 设置边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 填充数据
for product in products:
    row = [
        product['id'],
        product['name'],
        product['platform'],
        product['price'],
        product['publishTime'],
        product['url']
   ]
    ws.append(row)

# 设置数据行样式
data_font = Font(name='微软雅黑', size=10)
data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

for row in ws.iter_rows(min_row=2, max_row=len(products)+1):
    for cell in row:
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

# 调整列宽
ws.column_dimensions['A'].width = 8   # 序号
ws.column_dimensions['B'].width = 20  # 商品名称
ws.column_dimensions['C'].width = 10  # 平台
ws.column_dimensions['D'].width = 15  # 价格
ws.column_dimensions['E'].width = 18  # 发布时间
ws.column_dimensions['F'].width = 60  # 商品链接

# 保存文件
output_file = 'pxb7_products_50.xlsx'
wb.save(output_file)
print(f"Excel文件已生成: {output_file}")
print(f"共导出 {len(products)} 条商品数据")
