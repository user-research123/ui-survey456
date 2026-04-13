#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
七麦数据iOS畅销榜爬虫 - API方式
使用提供的Cookie通过API接口获取榜单数据
"""

import requests
import json
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io

# 配置
COOKIE_STRING = "qm_check=A1sdRUIQChtxen8pI0dAOQkKWVIeEHh+c3QgRioNDBgWFWVXXl1VRl0XXEcpCAkWUBd/BBlgRldJRjIGCwkfVl5UWVxUFG4AFBQBFxdTFxsQU1FVV1NHXEVYVElWBRsCHAkSSQ%3D%3D; gr_user_id=c6a9f667-62b3-4eba-b3a4-b072ace1ce68; ada35577182650f1_gr_last_sent_cs1=qm6956291145; AUTHKEY=YObS%2Bb5aF6FkrQTv71fWfBW7R%2FbXpK2zgfR%2FZmOs2eyv0AaBm9hp4iMLprQBImO9gSqWvh6Q1fJKxTiL%2FljgRkdeTQSz3F7h8hnMC%2BWNnnTPGmKP5UyYRQ%3D%3D; USERINFO=SjwgrvU8DAc5wQicVWjjsUlRdswW1BhyLq4j2G5GMvg3teKZIi2nJoiwraWA07DwoPpKFazjbsIgkBKMwE6HQ8lxmY5LPgIGaKbK%2FItqz0OeWk%2F%2FldCTsG%2BRAqMlcyjC7LlFz0SEuSmrXpiVqFfo6Q%3D%3D; PHPSESSID=f5imn4jn0mu4a8gilt0agmuod8; ada35577182650f1_gr_session_id=a691aa58-e0d3-4234-b79f-ec223eda031b; ada35577182650f1_gr_last_sent_sid_with_cs1=a691aa58-e0d3-4234-b79f-ec223eda031b; ada35577182650f1_gr_session_id_sent_vst=a691aa58-e0d3-4234-b79f-ec223eda031b; ada35577182650f1_gr_cs1=qm6956291145; synct=1775967610.773; syncd=-15"

# 解析Cookie为字典
def parse_cookies(cookie_string):
    cookies = {}
    for item in cookie_string.split('; '):
        if '=' in item:
            key, value = item.split('=', 1)
            cookies[key.strip()] = value.strip()
    return cookies

# 七麦数据API端点（需要根据实际情况调整）
# 注意：这些是基于常见模式的推测，实际可能需要从浏览器网络请求中获取
API_ENDPOINTS = {
    # 尝试不同的API端点
    'rank_list': 'https://api.qimai.cn/rank/index',
    'rank_index': 'https://api.qimai.cn/rankIndex',
}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Referer': 'https://www.qimai.cn/',
    'Origin': 'https://www.qimai.cn',
}

def fetch_rank_data():
    """获取榜单数据"""
    cookies = parse_cookies(COOKIE_STRING)
    
    # 参数配置 - iOS游戏畅销榜
    params = {
        'brand': 'all',      # 全部品牌
        'device': 'iphone',  # iPhone设备
        'country': 'cn',     # 中国区
        'genre': '6014',     # 游戏分类
        'date': datetime.now().strftime('%Y-%m-%d'),  # 当天日期
    }
    
    print(f"正在获取 {params['date']} 的iOS游戏畅销榜数据...")
    print(f"请求参数: {params}")
    
    # 尝试多个可能的API端点
    for endpoint_name, base_url in API_ENDPOINTS.items():
        try:
            print(f"\n尝试API端点: {endpoint_name}")
            
            # 构造完整URL
            if endpoint_name == 'rank_list':
                url = base_url
                response = requests.get(
                    url, 
                    params=params,
                    headers=HEADERS,
                    cookies=cookies,
                    timeout=10
                )
            else:
                # 其他端点可能需要不同的参数格式
                continue
            
            print(f"状态码: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                print(f"响应数据预览: {json.dumps(data, ensure_ascii=False)[:500]}")
                
                # 检查是否成功获取数据
                if data and isinstance(data, dict):
                    return data
                    
        except Exception as e:
            print(f"请求失败: {str(e)}")
            continue
    
    return None

def parse_rank_data(raw_data):
    """解析榜单数据"""
    if not raw_data:
        return []
    
    games = []
    
    # 根据实际API响应结构调整解析逻辑
    # 这里提供几种常见的数据结构解析方式
    
    # 方式1: 直接列表结构
    if isinstance(raw_data, list):
        for item in raw_data[:20]:  # 只取前20名
            game = {
                'rank': item.get('rank', item.get('index', '')),
                'name': item.get('appName', item.get('name', item.get('title', ''))),
                'developer': item.get('companyName', item.get('developer', item.get('publisher', ''))),
            }
            games.append(game)
    
    # 方式2: 嵌套在某个键下
    elif isinstance(raw_data, dict):
        # 尝试常见的数据键名
        data_keys = ['appList', 'list', 'data', 'result', 'apps', 'rankList']
        
        for key in data_keys:
            if key in raw_data:
                items = raw_data[key]
                if isinstance(items, list):
                    for item in items[:20]:
                        game = {
                            'rank': item.get('rank', item.get('index', '')),
                            'name': item.get('appName', item.get('name', item.get('title', ''))),
                            'developer': item.get('companyName', item.get('developer', item.get('publisher', ''))),
                        }
                        games.append(game)
                    break
        
        # 如果还没找到，尝试遍历所有键
        if not games:
            for key, value in raw_data.items():
                if isinstance(value, list) and len(value) > 0:
                    for item in value[:20]:
                        if isinstance(item, dict):
                            game = {
                                'rank': item.get('rank', item.get('index', '')),
                                'name': item.get('appName', item.get('name', item.get('title', ''))),
                                'developer': item.get('companyName', item.get('developer', item.get('publisher', ''))),
                            }
                            games.append(game)
                    if games:
                        break
    
    return games

def save_to_excel(games, filename='ios_grossing_rank.xlsx'):
    """保存为Excel文件"""
    if not games:
        print("没有数据可保存")
        return
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "iOS畅销榜TOP20"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ['排名', '游戏名称', '开发商/发行商']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入数据
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row_num, game in enumerate(games, 2):
        ws.cell(row=row_num, column=1, value=game['rank']).alignment = data_alignment
        ws.cell(row=row_num, column=2, value=game['name']).alignment = data_alignment
        ws.cell(row=row_num, column=3, value=game['developer']).alignment = data_alignment
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    
    # 保存文件（添加UTF-8 BOM以确保Excel兼容）
    wb.save(filename)
    print(f"\n数据已保存到: {filename}")
    print(f"共保存 {len(games)} 条记录")

def main():
    print("=" * 60)
    print("七麦数据 iOS游戏畅销榜爬虫")
    print("=" * 60)
    
    # 获取数据
    raw_data = fetch_rank_data()
    
    if not raw_data:
        print("\n❌ 未能通过API获取数据")
        print("\n可能的原因:")
        print("1. Cookie已过期或无效")
        print("2. API端点不正确")
        print("3. 需要额外的认证参数（如analysis、synct等）")
        print("\n建议:")
        print("- 在浏览器中打开 https://www.qimai.cn/rank/index/brand/all/device/iphone/country/cn/genre/6014")
        print("- 按F12打开开发者工具，切换到Network标签")
        print("- 刷新页面，查找包含'rank'或'list'的API请求")
        print("- 复制完整的请求URL和Headers中的Cookie")
        return
    
    # 解析数据
    games = parse_rank_data(raw_data)
    
    if not games:
        print("\n⚠️ 未能解析出有效的榜单数据")
        print(f"原始数据结构: {type(raw_data)}")
        print(f"原始数据内容: {json.dumps(raw_data, ensure_ascii=False)[:1000]}")
        return
    
    # 显示结果
    print(f"\n✅ 成功获取 {len(games)} 条榜单数据:\n")
    print(f"{'排名':<6} {'游戏名称':<30} {'开发商'}")
    print("-" * 70)
    for game in games:
        print(f"{game['rank']:<6} {game['name']:<30} {game['developer']}")
    
    # 保存到Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ios_grossing_rank_{timestamp}.xlsx'
    save_to_excel(games, filename)

if __name__ == '__main__':
    main()
