#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建示例Excel文件，展示七麦数据榜单的标准格式
注意：这是示例数据，实际使用时需要通过API或浏览器自动化获取真实数据
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def create_sample_excel():
    """创建示例Excel文件"""
    
    # 示例数据（基于常见的iOS游戏畅销榜）
    sample_games = [
        {'rank': 1, 'name': '王者荣耀', 'developer': '腾讯科技'},
        {'rank': 2, 'name': '和平精英', 'developer': '腾讯科技'},
        {'rank': 3, 'name': '原神', 'developer': 'miHoYo Limited'},
        {'rank': 4, 'name': '梦幻西游', 'developer': '网易移动游戏'},
        {'rank': 5, 'name': '三国志·战略版', 'developer': 'Lingxi Games Inc.'},
        {'rank': 6, 'name': '率土之滨', 'developer': '网易移动游戏'},
        {'rank': 7, 'name': '阴阳师', 'developer': '网易移动游戏'},
        {'rank': 8, 'name': '火影忍者', 'developer': '腾讯科技'},
        {'rank': 9, 'name': '完美世界', 'developer': 'Perfect World Co., Ltd'},
        {'rank': 10, 'name': '明日之后', 'developer': '网易移动游戏'},
        {'rank': 11, 'name': '问道', 'developer': 'G-bits Network Technology'},
        {'rank': 12, 'name': '天龙八部手游', 'developer': '腾讯科技'},
        {'rank': 13, 'name': '倩女幽魂', 'developer': '网易移动游戏'},
        {'rank': 14, 'name': '乱世王者', 'developer': '腾讯科技'},
        {'rank': 15, 'name': '剑与远征', 'developer': 'Lilith Games'},
        {'rank': 16, 'name': '万国觉醒', 'developer': 'Lilith Games'},
        {'rank': 17, 'name': '新笑傲江湖', 'developer': 'Perfect World Co., Ltd'},
        {'rank': 18, 'name': '最强蜗牛', 'developer': 'X.D. Network Inc.'},
        {'rank': 19, 'name': '荒野乱斗', 'developer': 'Supercell Oy'},
        {'rank': 20, 'name': '第五人格', 'developer': '网易移动游戏'},
    ]
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "iOS游戏畅销榜TOP20"
    
    # 添加抓取信息
    ws.cell(row=1, column=1, value=f"抓取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=2, column=1, value=f"数据来源: 七麦数据 (qimai.cn)")
    ws.cell(row=3, column=1, value=f"榜单类型: iOS中国区 - 游戏畅销榜")
    
    # 空一行
    row_offset = 5
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12, name='微软雅黑')
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ['排名', '游戏名称', '开发商/发行商']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_offset, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 数据样式
    data_font = Font(size=11, name='微软雅黑')
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 交替行颜色
    light_blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # 写入数据
    for idx, game in enumerate(sample_games):
        row_num = row_offset + 1 + idx
        
        # 排名
        rank_cell = ws.cell(row=row_num, column=1, value=game['rank'])
        rank_cell.font = data_font
        rank_cell.alignment = Alignment(horizontal="center", vertical="center")
        rank_cell.fill = light_blue_fill if idx % 2 == 0 else white_fill
        
        # 游戏名称
        name_cell = ws.cell(row=row_num, column=2, value=game['name'])
        name_cell.font = data_font
        name_cell.alignment = data_alignment
        name_cell.fill = light_blue_fill if idx % 2 == 0 else white_fill
        
        # 开发商
        dev_cell = ws.cell(row=row_num, column=3, value=game['developer'])
        dev_cell.font = data_font
        dev_cell.alignment = data_alignment
        dev_cell.fill = light_blue_fill if idx % 2 == 0 else white_fill
    
    # 调整列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 45
    
    # 添加边框
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=row_offset, max_row=row_offset + len(sample_games), 
                           min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # 保存文件（openpyxl默认会添加UTF-8 BOM）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ios_grossing_rank_sample_{timestamp}.xlsx'
    wb.save(filename)
    
    print(f"✅ 示例Excel文件已创建: {filename}")
    print(f"📊 包含 {len(sample_games)} 条示例数据")
    print(f"\n⚠️  注意: 这是示例数据，用于展示Excel格式")
    print(f"   要获取真实数据，请运行 qimai_crawler_v2.py 并配置有效的Cookie")
    
    return filename

if __name__ == '__main__':
    create_sample_excel()
