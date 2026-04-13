#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
七麦数据iOS畅销榜爬虫 v2
通过分析浏览器网络请求，使用正确的API端点获取榜单数据
"""

import requests
import json
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import base64
import hashlib

# 用户提供的完整Cookie
COOKIE_STRING = "qm_check=A1sdRUIQChtxen8pI0dAOQkKWVIeEHh+c3QgRioNDBgWFWVXXl1VRl0XXEcpCAkWUBd/BBlgRldJRjIGCwkfVl5UWVxUFG4AFBQBFxdTFxsQU1FVV1NHXEVYVElWBRsCHAkSSQ%3D%3D; gr_user_id=c6a9f667-62b3-4eba-b3a4-b072ace1ce68; ada35577182650f1_gr_last_sent_cs1=qm6956291145; AUTHKEY=YObS%2Bb5aF6FkrQTv71fWfBW7R%2FbXpK2zgfR%2FZmOs2eyv0AaBm9hp4iMLprQBImO9gSqWvh6Q1fJKxTiL%2FljgRkdeTQSz3F7h8hnMC%2BWNnnTPGmKP5UyYRQ%3D%3D; USERINFO=SjwgrvU8DAc5wQicVWjjsUlRdswW1BhyLq4j2G5GMvg3teKZIi2nJoiwraWA07DwoPpKFazjbsIgkBKMwE6HQ8lxmY5LPgIGaKbK%2FItqz0OeWk%2F%2FldCTsG%2BRAqMlcyjC7LlFz0SEuSmrXpiVqFfo6Q%3D%3D; PHPSESSID=f5imn4jn0mu4a8gilt0agmuod8; ada35577182650f1_gr_session_id=a691aa58-e0d3-4234-b79f-ec223eda031b; ada35577182650f1_gr_last_sent_sid_with_cs1=a691aa58-e0d3-4234-b79f-ec223eda031b; ada35577182650f1_gr_session_id_sent_vst=a691aa58-e0d3-4234-b79f-ec223eda031b; ada35577182650f1_gr_cs1=qm6956291145; synct=1775967610.773; syncd=-15"

def parse_cookies(cookie_string):
    """解析Cookie字符串为字典"""
    cookies = {}
    for item in cookie_string.split('; '):
        if '=' in item:
            key, value = item.split('=', 1)
            cookies[key.strip()] = value.strip()
    return cookies

def generate_analysis_param(params_dict):
    """
    生成七麦数据API所需的analysis参数
    这是七麦数据的签名机制，需要将参数进行特定编码
    """
    # 将参数字典转换为查询字符串
    query_string = '&'.join([f"{k}={v}" for k, v in sorted(params_dict.items())])
    
    # 七麦的analysis参数通常是base64编码的参数串
    # 这里简化处理，实际可能需要更复杂的加密逻辑
    encoded = base64.b64encode(query_string.encode()).decode()
    return encoded

def fetch_rank_data_v2():
    """使用正确的API端点获取榜单数据"""
    cookies = parse_cookies(COOKIE_STRING)
    
    # 根据浏览器观察到的API模式，尝试不同的端点
    # 七麦数据的API通常需要analysis参数
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'Referer': 'https://www.qimai.cn/rank/index/brand/all/device/iphone/country/cn/genre/6014',
        'Origin': 'https://www.qimai.cn',
        'Connection': 'keep-alive',
    }
    
    # 尝试方法1: 直接使用rank/index API（之前试过，返回空数据）
    print("方法1: 尝试 rank/index API...")
    params1 = {
        'brand': 'all',
        'device': 'iphone',
        'country': 'cn',
        'genre': '6014',
        'date': datetime.now().strftime('%Y-%m-%d'),
    }
    
    try:
        response1 = requests.get(
            'https://api.qimai.cn/rank/index',
            params=params1,
            headers=headers,
            cookies=cookies,
            timeout=10
        )
        print(f"状态码: {response1.status_code}")
        data1 = response1.json()
        print(f"响应: {json.dumps(data1, ensure_ascii=False)[:300]}")
        
        # 检查是否有数据
        if data1.get('rankInfo') or data1.get('appList') or data1.get('list'):
            print("✅ 方法1成功获取数据")
            return data1
    except Exception as e:
        print(f"方法1失败: {e}")
    
    # 尝试方法2: 使用userRequest API（从浏览器网络请求中观察到）
    print("\n方法2: 尝试 userRequest/index API...")
    try:
        # 这个API可能需要POST请求和特定的analysis参数
        params2 = {
            'analysis': 'dkZJEhsdCyhQQEIGFxUWXwsCAhA4WksEBQBXV1cIBVRSVChbSg%3D%3D',  # 从浏览器捕获的值
        }
        
        response2 = requests.post(
            'https://api.qimai.cn/userRequest/index',
            params=params2,
            headers=headers,
            cookies=cookies,
            timeout=10
        )
        print(f"状态码: {response2.status_code}")
        data2 = response2.json()
        print(f"响应预览: {json.dumps(data2, ensure_ascii=False)[:500]}")
        
        if data2 and isinstance(data2, dict):
            print("✅ 方法2可能成功")
            return data2
    except Exception as e:
        print(f"方法2失败: {e}")
    
    # 尝试方法3: 使用index/index API（也从浏览器观察到）
    print("\n方法3: 尝试 index/index API...")
    try:
        params3 = {
            'analysis': 'dkZJDgYcHAIaWFkHARl5FVRXU19MSkwFCQZaUSEaBQ%3D%3D',
        }
        
        response3 = requests.get(
            'https://api.qimai.cn/index/index',
            params=params3,
            headers=headers,
            cookies=cookies,
            timeout=10
        )
        print(f"状态码: {response3.status_code}")
        data3 = response3.json()
        print(f"响应预览: {json.dumps(data3, ensure_ascii=False)[:500]}")
        
        if data3 and isinstance(data3, dict):
            print("✅ 方法3可能成功")
            return data3
    except Exception as e:
        print(f"方法3失败: {e}")
    
    return None

def extract_games_from_response(raw_data):
    """从API响应中提取游戏列表"""
    if not raw_data:
        return []
    
    games = []
    
    # 递归查找包含游戏信息的列表
    def find_game_list(obj, depth=0):
        if depth > 5:  # 防止无限递归
            return []
        
        if isinstance(obj, list):
            # 检查这个列表是否包含游戏数据
            if len(obj) > 0 and isinstance(obj[0], dict):
                first_item = obj[0]
                # 检查是否包含游戏相关字段
                if any(key in first_item for key in ['appName', 'name', 'rank', 'companyName']):
                    return obj
        
        if isinstance(obj, dict):
            for key, value in obj.items():
                result = find_game_list(value, depth + 1)
                if result:
                    return result
        
        return []
    
    game_list = find_game_list(raw_data)
    
    if not game_list:
        print("⚠️ 未找到游戏列表数据")
        print(f"完整响应结构: {json.dumps(raw_data, ensure_ascii=False, indent=2)[:2000]}")
        return []
    
    # 提取前20条记录
    for item in game_list[:20]:
        if not isinstance(item, dict):
            continue
            
        game = {
            'rank': item.get('rankIndex', item.get('rank', item.get('index', ''))),
            'name': item.get('appName', item.get('name', item.get('title', ''))),
            'developer': item.get('companyName', item.get('developer', item.get('publisher', ''))),
        }
        
        # 只添加有有效名称的游戏
        if game['name']:
            games.append(game)
    
    return games

def save_to_excel(games, filename='ios_grossing_rank.xlsx'):
    """保存为Excel文件（带UTF-8 BOM）"""
    if not games:
        print("没有数据可保存")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "iOS畅销榜TOP20"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ['排名', '游戏名称', '开发商/发行商']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入数据
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row_num, game in enumerate(games, 2):
        ws.cell(row=row_num, column=1, value=game['rank']).alignment = data_alignment
        ws.cell(row=row_num, column=2, value=game['name']).alignment = data_alignment
        ws.cell(row=row_num, column=3, value=game['developer']).alignment = data_alignment
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    
    # 保存文件
    wb.save(filename)
    print(f"\n✅ 数据已保存到: {filename}")
    print(f"共保存 {len(games)} 条记录")

def main():
    print("=" * 70)
    print("七麦数据 iOS游戏畅销榜爬虫 v2")
    print(f"抓取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    
    # 获取数据
    raw_data = fetch_rank_data_v2()
    
    if not raw_data:
        print("\n❌ 所有API尝试均失败")
        print("\n可能的解决方案:")
        print("1. Cookie可能已过期，请重新从浏览器获取最新Cookie")
        print("2. analysis参数可能需要动态生成")
        print("3. 可能需要在浏览器中手动登录后复制完整的请求")
        return
    
    # 提取游戏数据
    games = extract_games_from_response(raw_data)
    
    if not games:
        print("\n❌ 未能从响应中提取游戏数据")
        return
    
    # 显示结果
    print(f"\n{'='*70}")
    print(f"✅ 成功获取 {len(games)} 条榜单数据:\n")
    print(f"{'排名':<6} {'游戏名称':<35} {'开发商'}")
    print("-" * 70)
    for game in games:
        print(f"{str(game['rank']):<6} {game['name']:<35} {game['developer']}")
    
    # 保存到Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ios_grossing_rank_{timestamp}.xlsx'
    save_to_excel(games, filename)
    
    print(f"\n{'='*70}")
    print("爬虫执行完成！")

if __name__ == '__main__':
    main()
