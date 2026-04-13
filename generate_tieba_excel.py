#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成百度贴吧帖子数据Excel文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 贴吧基本信息
tieba_info = {
    "贴吧名称": "洛克王国世界吧",
    "关注人数": 7931,
    "贴子总数": "5.6W",
    "数据来源": "浏览器自动化爬取",
    "爬取时间": "2026-03-25 13:42"
}

# 帖子数据(从readability提取)
posts_data = [
    {
        "序号": 1,
        "作者": "吟月悲",
        "作者等级": "高级粉丝",
        "发布时间": "1分钟前",
        "标题": "明天洛克王国公测上线有没有一起玩的滴滴本人纯新手拉了个小群...",
        "内容摘要": "明天洛克王国公测上线有没有一起玩的滴滴本人纯新手拉了个小群，有爱讨论爱聊天的内测大佬可以进分享游戏心得，也欢迎跟我一样的小白，刚建人不多欢迎欢迎 #晒出你的极品id#",
        "分享数": 7,
        "回复数": 1
    },
    {
        "序号": 2,
        "作者": "加糖的冰咖啡",
        "作者等级": "初级粉丝",
        "发布时间": "3分钟前",
        "标题": "移动端是不是要比pc端少很多东西啊",
        "内容摘要": "移动端是不是要比pc端少很多东西啊",
        "分享数": 3,
        "回复数": 0
    },
    {
        "序号": 3,
        "作者": "G.E.M.歌莉雅",
        "作者等级": "高级粉丝",
        "发布时间": "4分钟前",
        "标题": "心碎了家人,大家看到好的号码就不要再刷了为了个所谓的\"超靓\"我失去了太多……",
        "内容摘要": "心碎了家人,大家看到好的号码就不要再刷了为了个所谓的\"超靓\"我失去了太多……第一张图是我过程中遇到的\"幸运号码\"，最后为了搏\"超靓\"，给我把这个刷掉了。",
        "分享数": 23,
        "回复数": 2
    }
]

# 创建工作簿
wb = Workbook()

# 创建帖子列表工作表
ws_posts = wb.active
ws_posts.title = "帖子列表"

# 设置表头
headers = ["序号", "作者", "作者等级", "发布时间", "标题", "内容摘要", "分享数", "回复数"]
ws_posts.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws_posts.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 填充数据
for post in posts_data:
    row_data = [
        post["序号"],
        post["作者"],
        post["作者等级"],
        post["发布时间"],
        post["标题"],
        post["内容摘要"],
        post["分享数"],
        post["回复数"]
    ]
    ws_posts.append(row_data)

# 设置列宽
column_widths = [8, 15, 12, 12, 40, 60, 10,10]
for i, width in enumerate(column_widths, 1):
    ws_posts.column_dimensions[get_column_letter(i)].width = width

# 设置数据行对齐方式
for row in ws_posts.iter_rows(min_row=2, max_row=len(posts_data)+1):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# 创建贴吧信息工作表
ws_info = wb.create_sheet(title="贴吧信息")

#填充贴吧信息
info_data = [
    ["属性", "值"],
    ["贴吧名称", tieba_info["贴吧名称"]],
    ["关注人数", tieba_info["关注人数"]],
    ["贴子总数", tieba_info["贴子总数"]],
    ["数据来源", tieba_info["数据来源"]],
    ["爬取时间", tieba_info["爬取时间"]]
]

for row_data in info_data:
    ws_info.append(row_data)

# 设置贴吧信息工作表的列宽
ws_info.column_dimensions['A'].width = 15
ws_info.column_dimensions['B'].width = 40

# 设置表头样式
for col_num in range(1, 3):
    cell = ws_info.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 保存文件
output_file = "洛克王国世界吧_帖子列表_完整版.xlsx"
wb.save(output_file)
print(f"Excel文件已生成: {output_file}")
print(f"共提取 {len(posts_data)} 条帖子数据")