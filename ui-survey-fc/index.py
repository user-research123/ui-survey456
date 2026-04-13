# -*- coding: utf-8 -*-
"""
阿里云函数计算 - UI 调研问卷后端处理器
支持 HTTP 触发器，接收问卷提交数据并存储
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
import requests


def load_or_create_excel():
    """加载或创建 Excel 文件"""
    excel_path = '/tmp/survey_data.xlsx'
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "问卷数据"
        
        # 表头（根据你的 12 道题调整）
        headers = [
            '提交时间', '用户 ID', 'Q1_游戏经验', 'Q2_UI 风格偏好', 
            'Q3_布局偏好', 'Q4_色彩偏好', 'Q5_字体大小', 'Q6_按钮样式',
            'Q7_交互动效', 'Q8_原型评分', 'Q9_满意度', 'Q10_改进建议',
            'Q11_设备类型', 'Q12_其他反馈'
        ]
        ws.append(headers)
        wb.save(excel_path)
    
    return excel_path


def save_to_excel(data):
    """保存数据到 Excel"""
    excel_path = load_or_create_excel()
    
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 将表单数据转换为行
    row = [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        data.get('user_id', 'anonymous'),
        data.get('q1_game_exp', ''),
        data.get('q2_style', ''),
        data.get('q3_layout', ''),
        data.get('q4_color', ''),
        data.get('q5_font', ''),
        data.get('q6_button', ''),
        data.get('q7_animation', ''),
        data.get('q8_prototype_rate', ''),
        data.get('q9_satisfaction', ''),
        data.get('q10_suggestion', ''),
        data.get('q11_device', ''),
        data.get('q12_other', '')
    ]
    ws.append(row)
    wb.save(excel_path)
    
    return True


def send_to_dingtalk(data, submit_time):
    """发送提交通知到钉钉机器人（可选）"""
    webhook_url = os.environ.get('DINGTALK_WEBHOOK')
    
    if not webhook_url:
        return False
    
    # 构建通知消息
    content = f"""📊 UI 调研问卷新提交
提交时间：{submit_time}
用户设备：{data.get('q11_device', '未知')}
满意度评分：{data.get('q9_satisfaction', '未填写')}
改进建议：{data.get('q10_suggestion', '无')[:50]}..."""
    
    payload = {
        "msgtype": "text",
        "text": {
            "content": content
        }
    }
    
    try:
        response = requests.post(
            webhook_url,
            json=payload,
            headers={'Content-Type': 'application/json'},
            timeout=5
        )
        return response.status_code == 200
    except Exception as e:
        print(f"钉钉推送失败：{str(e)}")
        return False


def handler(environ, start_response):
    """
    阿里云函数计算 HTTP 触发器入口函数
    兼容 Flask/Werkzeug 请求格式
    """
    try:
        # 获取请求方法
        request_method = environ.get('REQUEST_METHOD', 'GET')
        
        # CORS 预检请求处理
        if request_method == 'OPTIONS':
            headers = [
                ('Content-Type', 'application/json'),
                ('Access-Control-Allow-Origin', '*'),
                ('Access-Control-Allow-Methods', 'GET, POST, OPTIONS'),
                ('Access-Control-Allow-Headers', 'Content-Type')
            ]
            start_response('200 OK', headers)
            return [b'']
        
        # 处理 GET 请求（健康检查）
        if request_method == 'GET':
            response_body = {
                'status': 'ok',
                'message': 'UI Survey API is running',
                'timestamp': datetime.now().isoformat()
            }
            
            headers = [
                ('Content-Type', 'application/json'),
                ('Access-Control-Allow-Origin', '*')
            ]
            start_response('200 OK', headers)
            return [json.dumps(response_body).encode('utf-8')]
        
        # 处理 POST 请求（接收问卷提交）
        if request_method == 'POST':
            # 读取请求体
            try:
                request_body_size = int(environ.get('CONTENT_LENGTH', 0))
            except (ValueError):
                request_body_size = 0
            
            body = environ['wsgi.input'].read(request_body_size)
            data = json.loads(body.decode('utf-8'))
            
            # 保存数据到 Excel
            submit_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            save_to_excel(data)
            
            # 发送到钉钉（可选）
            send_to_dingtalk(data, submit_time)
            
            # 返回成功响应
            response_body = {
                'status': 'success',
                'message': '问卷提交成功',
                'submit_time': submit_time
            }
            
            headers = [
                ('Content-Type', 'application/json'),
                ('Access-Control-Allow-Origin', '*')
            ]
            start_response('200 OK', headers)
            return [json.dumps(response_body).encode('utf-8')]
        
        # 不支持的请求方法
        headers = [('Content-Type', 'application/json')]
        start_response('405 Method Not Allowed', headers)
        return [json.dumps({'error': 'Method not allowed'}).encode('utf-8')]
    
    except Exception as e:
        # 错误处理
        error_response = {
            'status': 'error',
            'message': str(e)
        }
        headers = [
            ('Content-Type', 'application/json'),
            ('Access-Control-Allow-Origin', '*')
        ]
        start_response('500 Internal Server Error', headers)
        return [json.dumps(error_response).encode('utf-8')]
