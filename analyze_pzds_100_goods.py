#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析盼之网站前100个商品数据
"""

import json
import statistics
from collections import Counter

# 商品数据（从浏览器提取的100个商品）
goods_data = [
    {"id": 1, "title": "王者荣耀：世界火热招商中！！", "description": "王者荣耀：世界火热招商中！！", "platform": "", "price": 999999},
    {"id": 2, "title": "元流返利·觉醒开局 返利账号上架专场", "description": "元流返利·觉醒开局 返利账号上架专场", "platform": "", "price": 9999999},
    {"id": 3, "title": "单字id胆，常见单字，基本都认识，诚心出售", "description": "单字id胆，常见单字，基本都认识，诚心出售", "platform": "", "price": 1000},
    {"id": 4, "title": "极品二字id：玉佩", "description": "极品二字id：玉佩", "platform": "", "price": 720},
    {"id": 5, "title": "ID：杨颖baby   喜欢的直接拍", "description": "ID：杨颖baby   喜欢的直接拍", "platform": "", "price": 800},
    {"id": 6, "title": "LGBT系极品词组：小受", "description": "LGBT系极品词组：小受", "platform": "", "price": 888},
    {"id": 7, "title": "ID衿", "description": "ID衿", "platform": "", "price": 999},
    {"id": 8, "title": "极品帅哥id:偏爱纵容，喜欢可谈，有搭配情侣id女号:性格娇纵", "description": "极品帅哥id:偏爱纵容，喜欢可谈，有搭配情侣id女号:性格娇纵", "platform": "", "price": 900},
    {"id": 9, "title": "极品双子ID：绘织", "description": "极品双子ID：绘织", "platform": "", "price": 1200},
    {"id": 10, "title": "双字词组🆔推辞", "description": "双字词组🆔推辞", "platform": "", "price": 1100},
    {"id": 11, "title": "双字🆔：好宅", "description": "双字🆔：好宅", "platform": "", "price": 799},
    {"id": 12, "title": "极品单字🆔瞻", "description": "极品单字🆔瞻", "platform": "", "price": 1000},
    {"id": 13, "title": "出双字id星影", "description": "出双字id星影", "platform": "", "price": 999},
    {"id": 14, "title": "王者荣耀世界id：铁杉（非强组，一个树的名字）", "description": "王者荣耀世界id：铁杉（非强组，一个树的名字）", "platform": "", "price": 999},
    {"id": 15, "title": "双字id【瞳孔】", "description": "双字id【瞳孔】", "platform": "", "price": 888},
    {"id": 16, "title": "极品双字ID：绣织", "description": "极品双字ID：绣织", "platform": "", "price": 1000},
    {"id": 17, "title": "正版id婚戒", "description": "正版id婚戒", "platform": "", "price": 888},
    {"id": 18, "title": "极品美女id:性格娇纵，喜欢可谈，有搭配情侣id男号:偏爱纵容，一起购买可更优惠", "description": "极品美女id:性格娇纵，喜欢可谈，有搭配情侣id男号:偏爱纵容，一起购买可更优惠", "platform": "", "price": 900},
    {"id": 19, "title": "Id：愿中国永无难，喜欢的可以聊", "description": "Id：愿中国永无难，喜欢的可以聊", "platform": "", "price": 1000},
    {"id": 20, "title": "出极品单子id实", "description": "出极品单子id实", "platform": "", "price": 888},
    {"id": 21, "title": "极品双字id色骨", "description": "极品双字id色骨", "platform": "", "price": 888},
    {"id": 22, "title": "极品昵称，吾乃吴彦祖", "description": "极品昵称，吾乃吴彦祖", "platform": "", "price": 888},
    {"id": 23, "title": "极品id", "description": "极品id", "platform": "", "price": 888},
    {"id": 24, "title": "游戏账号", "description": "游戏账号", "platform": "", "price": 888},
    {"id": 25, "title": "极品昵称，刘亦菲lyf", "description": "极品昵称，刘亦菲lyf", "platform": "", "price": 1288},
    {"id": 26, "title": "极品id", "description": "极品id", "platform": "", "price": 888},
    {"id": 27, "title": "双字Id严父", "description": "双字Id严父", "platform": "", "price": 1169},
    {"id": 28, "title": "极品昵称，想你的声音", "description": "极品昵称，想你的声音", "platform": "", "price": 888},
    {"id": 29, "title": "极品id", "description": "极品id", "platform": "", "price": 888},
    {"id": 30, "title": "极品昵称，吾乃彭于晏", "description": "极品昵称，吾乃彭于晏", "platform": "", "price": 888},
    {"id": 31, "title": "999999999九个九绝版孤品名字", "description": "999999999九个九绝版孤品名字", "platform": "", "price": 199999},
    {"id": 32, "title": "王者荣耀世界id", "description": "王者荣耀世界id", "platform": "", "price": 101},
    {"id": 33, "title": "王者荣耀世界极品单字id：愮", "description": "王者荣耀世界极品单字id：愮", "platform": "", "price": 188888},
    {"id": 34, "title": "id666666", "description": "id666666", "platform": "", "price": 20000},
    {"id": 35, "title": "单字id缙,诚心出售", "description": "单字id缙,诚心出售", "platform": "", "price": 500},
    {"id": 36, "title": "极品id：万瑶之王", "description": "极品id：万瑶之王", "platform": "", "price": 188888},
    {"id": 37, "title": "iD：茜姐", "description": "iD：茜姐", "platform": "", "price": 351},
    {"id": 38, "title": "王者荣耀世界极品二字ID", "description": "王者荣耀世界极品二字ID", "platform": "", "price": 298},
    {"id": 39, "title": "单字id：胡", "description": "单字id：胡", "platform": "", "price": 4111},
    {"id": 40, "title": "女用id ：甜雨眠", "description": "女用id ：甜雨眠", "platform": "", "price": 121},
    {"id": 41, "title": "安卓QQ\n¥ 357\n4小时前发布\n2人想要", "description": "安卓QQ ¥ 357 4小时前发布 2人想要", "platform": "安卓QQ", "price": 357},
    {"id": 42, "title": "王者荣耀世界ID:i女", "description": "王者荣耀世界ID:i女", "platform": "", "price": 350},
    {"id": 43, "title": "王者荣耀世界单字ID匍", "description": "王者荣耀世界单字ID匍", "platform": "", "price": 180},
    {"id": 44, "title": "单字：噛，连同QQ号岀售", "description": "单字：噛，连同QQ号岀售", "platform": "", "price": 199},
    {"id": 45, "title": "id号。残忍", "description": "id号。残忍", "platform": "", "price": 3000},
    {"id": 46, "title": "王者荣耀世界单字ID颅，取敌方首级之意", "description": "王者荣耀世界单字ID颅，取敌方首级之意", "platform": "", "price": 666},
    {"id": 47, "title": "王者荣耀世界双字极品ID:碎今", "description": "王者荣耀世界双字极品ID:碎今", "platform": "", "price": 125},
    {"id": 48, "title": "双字id乾陇", "description": "双字id乾陇", "platform": "", "price": 288},
    {"id": 49, "title": "白敬亭", "description": "白敬亭", "platform": "", "price": 10066},
    {"id": 50, "title": "王者荣耀世界双字极品ID:裁判", "description": "王者荣耀世界双字极品ID:裁判", "platform": "", "price": 238},
    {"id": 51, "title": "双字id慕榕", "description": "双字id慕榕", "platform": "", "price": 288},
    {"id": 52, "title": "王者荣耀世界ID市民", "description": "王者荣耀世界ID市民", "platform": "", "price": 13000},
    {"id": 53, "title": "单字id：狂", "description": "单字id：狂", "platform": "", "price": 10000},
    {"id": 54, "title": "双字id栢枫", "description": "双字id栢枫", "platform": "", "price": 288},
    {"id": 55, "title": "御姐甜妹专用极简ld：Fie", "description": "御姐甜妹专用极简ld：Fie", "platform": "", "price": 10000},
    {"id": 56, "title": "双字小词组🆔青筋", "description": "双字小词组🆔青筋", "platform": "", "price": 530},
    {"id": 57, "title": "王者极品id", "description": "王者极品id", "platform": "", "price": 9999},
    {"id": 58, "title": "一眼大美女极品ld:池夢鲤", "description": "一眼大美女极品ld:池夢鲤", "platform": "", "price": 10000},
    {"id": 59, "title": "ID：张艺兴，无空白字符", "description": "ID：张艺兴，无空白字符", "platform": "", "price": 131452},
    {"id": 60, "title": "顶级ID宇宙，要的来", "description": "顶级ID宇宙，要的来", "platform": "", "price": 2400},
    {"id": 61, "title": "王者荣耀世界id真龙", "description": "王者荣耀世界id真龙", "platform": "", "price": 22000},
    {"id": 62, "title": "ID和号", "description": "ID和号", "platform": "", "price": 200},
    {"id": 63, "title": "极品单字🆔榜", "description": "极品单字🆔榜", "platform": "", "price": 5555},
    {"id": 64, "title": "极品双字🆔：感叹", "description": "极品双字🆔：感叹", "platform": "", "price": 1499},
    {"id": 65, "title": "单子ID粧", "description": "单子ID粧", "platform": "", "price": 245},
    {"id": 66, "title": "单字id菊", "description": "单字id菊", "platform": "", "price": 4117},
    {"id": 67, "title": "虞书欣", "description": "虞书欣", "platform": "", "price": 13456},
    {"id": 68, "title": "极品单字id涬", "description": "极品单字id涬", "platform": "", "price": 489},
    {"id": 69, "title": "极品ID妞子 ，可讲价", "description": "极品ID妞子 ，可讲价", "platform": "", "price": 350},
    {"id": 70, "title": "极品id，便宜", "description": "极品id，便宜", "platform": "", "price": 69},
    {"id": 71, "title": "三字🆔自习室", "description": "三字🆔自习室", "platform": "", "price": 300},
    {"id": 72, "title": "女生用的漂亮id", "description": "女生用的漂亮id", "platform": "", "price": 1500},
    {"id": 73, "title": "极品ID", "description": "极品ID", "platform": "", "price": 666},
    {"id": 74, "title": "极品二字id吟诗，无特殊符号", "description": "极品二字id吟诗，无特殊符号", "platform": "", "price": 520},
    {"id": 75, "title": "双ID败感", "description": "双ID败感", "platform": "", "price": 130},
    {"id": 76, "title": "两个一起出，可做闺蜜ID情侣ID，名字好听适合女生", "description": "两个一起出，可做闺蜜ID情侣ID，名字好听适合女生", "platform": "", "price": 520},
    {"id": 77, "title": "王者荣耀世界id听细雨绵绵", "description": "王者荣耀世界id听细雨绵绵", "platform": "", "price": 199},
    {"id": 78, "title": "极品单子ID佖", "description": "极品单子ID佖", "platform": "", "price": 188},
    {"id": 79, "title": "极品双字ID：绘猫", "description": "极品双字ID：绘猫", "platform": "", "price": 1800},
    {"id": 80, "title": "王者荣耀世界双字id溺晚", "description": "王者荣耀世界双字id溺晚", "platform": "", "price": 399},
    {"id": 81, "title": "王者荣耀世界极品单字id：矓", "description": "王者荣耀世界极品单字id：矓", "platform": "", "price": 188888},
    {"id": 82, "title": "王者荣耀世界ID：剫，要的来", "description": "王者荣耀世界ID：剫，要的来", "platform": "", "price": 120},
    {"id": 83, "title": "王者荣耀世界id夜色漫漫", "description": "王者荣耀世界id夜色漫漫", "platform": "", "price": 266},
    {"id": 84, "title": "安卓QQ\n¥ 69\n125人看过\n11人想要", "description": "安卓QQ ¥ 69 125人看过 11人想要", "platform": "安卓QQ", "price": 69},
    {"id": 85, "title": "单字id连q一起出", "description": "单字id连q一起出", "platform": "", "price": 266},
    {"id": 86, "title": "极品双字ID梁子", "description": "极品双字ID梁子", "platform": "", "price": 520},
    {"id": 87, "title": "ID   景甜  喜欢的价格可议", "description": "ID   景甜  喜欢的价格可议", "platform": "", "price": 40000},
    {"id": 88, "title": "单字id\n连q一起出", "description": "单字id\n连q一起出", "platform": "", "price": 366},
    {"id": 89, "title": "王者荣耀世界id:大肥婆西施", "description": "王者荣耀世界id:大肥婆西施", "platform": "", "price": 2778},
    {"id": 90, "title": "正版id花架", "description": "正版id花架", "platform": "", "price": 288},
    {"id": 91, "title": "单字id：埦", "description": "单字id：埦", "platform": "", "price": 311},
    {"id": 92, "title": "单字ID：嬀", "description": "单字ID：嬀", "platform": "", "price": 243},
    {"id": 93, "title": "王者荣耀世界正版id无代码", "description": "王者荣耀世界正版id无代码", "platform": "", "price": 2000},
    {"id": 94, "title": "便宜出出单字id唷", "description": "便宜出出单字id唷", "platform": "", "price": 400},
    {"id": 95, "title": "王者荣耀世界id", "description": "王者荣耀世界id", "platform": "", "price": 599},
    {"id": 96, "title": "双字ID：吻枫", "description": "双字ID：吻枫", "platform": "", "price": 288},
    {"id": 97, "title": "出极品双字id睡去", "description": "出极品双字id睡去", "platform": "", "price": 688},
    {"id": 98, "title": "王者世界ID：OOO", "description": "王者世界ID：OOO", "platform": "", "price": 100},
    {"id": 99, "title": "小极品带号出 lemo", "description": "小极品带号出 lemo", "platform": "", "price": 300},
    {"id": 100, "title": "王者荣耀世界ID：美女儿", "description": "王者荣耀世界ID：美女儿", "platform": "", "price": 130}
]

def categorize_goods(goods):
    """对商品进行分类"""
    categories = {
        '明星名人ID': [],
        '单字ID': [],
        '双字ID': [],
        '三字及以上ID': [],
        '特殊含义ID': [],
        '情侣/闺蜜ID': [],
        '招商/返利账号': [],
        '普通ID': []
    }
    
    for good in goods:
        title = good['title']
        desc = good['description'].lower()
        
        # 明星名人ID
        if any(name in title for name in ['杨颖', '吴彦祖', '刘亦菲', '彭于晏', '白敬亭', '张艺兴', '虞书欣', '景甜', 'baby']):
            categories['明星名人ID'].append(good)
        # 招商/返利账号
        elif '招商' in title or '返利' in title:
            categories['招商/返利账号'].append(good)
        # 单字ID
        elif '单字' in title or ('id' in desc and len([c for c in title if c.isalpha() or '\u4e00' <= c <= '\u9fff']) <= 3):
            categories['单字ID'].append(good)
        # 双字ID
        elif '双字' in title or '二字' in title:
            categories['双字ID'].append(good)
        # 三字及以上ID
        elif '三字' in title or any(word in title for word in ['自习室', '听细雨', '夜色漫漫', '大肥婆', '美女儿']):
            categories['三字及以上ID'].append(good)
        # 情侣/闺蜜ID
        elif '情侣' in title or '闺蜜' in title or '搭配' in title:
            categories['情侣/闺蜜ID'].append(good)
        # 特殊含义ID（包含特定词汇）
        elif any(word in title for word in ['极品', '绝版', '孤品', '顶级', '正版', '万瑶之王', '宇宙', '真龙', '狂']):
            categories['特殊含义ID'].append(good)
        else:
            categories['普通ID'].append(good)
    
    return categories

def analyze_price_distribution(prices):
    """分析价格分布"""
    if not prices:
        return {}
    
    # 价格区间统计
    ranges = {
        '0-500': 0,
        '500-1000': 0,
        '1000-5000': 0,
        '5000-10000': 0,
        '10000-50000': 0,
        '50000+': 0
    }
    
    for price in prices:
        if price < 500:
            ranges['0-500'] += 1
        elif price < 1000:
            ranges['500-1000'] += 1
        elif price < 5000:
            ranges['1000-5000'] += 1
        elif price < 10000:
            ranges['5000-10000'] += 1
        elif price < 50000:
            ranges['10000-50000'] += 1
        else:
            ranges['50000+'] += 1
    
    return ranges

def main():
    print("=" * 80)
    print("盼之网站前100个商品分析报告")
    print("=" * 80)
    
    # 提取价格列表
    prices = [good['price'] for good in goods_data if good['price'] is not None]
    
    # 商品分类
    categories = categorize_goods(goods_data)
    
    print("\n【商品类型分析】")
    print("-" * 80)
    print("商品类型多样化，主要包括以下类型商品：\n")
    
    for category, items in sorted(categories.items(), key=lambda x: len(x[1]), reverse=True):
        if items:
            percentage = len(items) / len(goods_data) * 100
            print(f"• {category}: {len(items)} 个 ({percentage:.1f}%)")
            # 显示该类别的部分示例
            for item in items[:3]:
                print(f"  - {item['title'][:50]}... (¥{item['price']:,})")
            if len(items) > 3:
                print(f"  ... 还有 {len(items) - 3} 个")
            print()
    
    # 价格分布分析
    print("\n【价格分布分析】")
    print("-" * 80)
    
    price_ranges = analyze_price_distribution(prices)
    total = len(prices)
    
    print(f"\n价格范围: ¥{min(prices):,} - ¥{max(prices):,}")
    print(f"平均价格: ¥{statistics.mean(prices):,.0f}")
    print(f"中位数价格: ¥{statistics.median(prices):,.0f}")
    
    high_price_count = sum(1 for p in prices if p >= 10000)
    print(f"高价商品(≥¥10,000): {high_price_count} 个 ({high_price_count/total*100:.1f}%)")
    
    print("\n价格区间分布:")
    for range_name, count in price_ranges.items():
        percentage = count / total * 100
        print(f"{range_name}: {count} 个 ({percentage:.1f}%)")
    
    # 生成CSV文件
    print("\n【生成数据文件】")
    print("-" * 80)
    
    import csv
    
    csv_file = '/Users/jiewen/.real/users/user-777da3823a68e71779b041c8e7807df0/workspace/pzds_100_goods.csv'
    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['序号', '商品标题', '商品描述', '平台', '价格(元)', '商品类型'])
        
        for good in goods_data:
            # 确定商品类型
            good_type = '普通ID'
            for category, items in categories.items():
                if good in items:
                    good_type = category
                    break
            
            writer.writerow([
                good['id'],
                good['title'].replace('\n', ' '),
                good['description'].replace('\n', ' '),
                good['platform'],
                good['price'],
                good_type
            ])
    
    print(f"✓ 已生成CSV文件: {csv_file}")
    
    # 生成Excel文件
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        excel_file = '/Users/jiewen/.real/users/user-777da3823a68e71779b041c8e7807df0/workspace/pzds_100_goods.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "商品列表"
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ['序号', '商品标题', '商品描述', '平台', '价格(元)', '商品类型']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for idx, good in enumerate(goods_data, 2):
            good_type = '普通ID'
            for category, items in categories.items():
                if good in items:
                    good_type = category
                    break
            
            ws.cell(row=idx, column=1, value=good['id'])
            ws.cell(row=idx, column=2, value=good['title'].replace('\n', ' '))
            ws.cell(row=idx, column=3, value=good['description'].replace('\n', ' '))
            ws.cell(row=idx, column=4, value=good['platform'])
            ws.cell(row=idx, column=5, value=good['price'])
            ws.cell(row=idx, column=6, value=good_type)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        # 创建统计分析sheet
        ws2 = wb.create_sheet("统计分析")
        
        # 写入统计信息
        ws2.cell(row=1, column=1, value="盼之网站前100个商品统计分析").font = Font(bold=True, size=14)
        
        row = 3
        ws2.cell(row=row, column=1, value="一、商品类型分布").font = Font(bold=True)
        row += 1
        for category, items in sorted(categories.items(), key=lambda x: len(x[1]), reverse=True):
            if items:
                percentage = len(items) / len(goods_data) * 100
                ws2.cell(row=row, column=1, value=f"{category}")
                ws2.cell(row=row, column=2, value=len(items))
                ws2.cell(row=row, column=3, value=f"{percentage:.1f}%")
                row += 1
        
        row += 2
        ws2.cell(row=row, column=1, value="二、价格分布分析").font = Font(bold=True)
        row += 1
        ws2.cell(row=row, column=1, value="价格范围")
        ws2.cell(row=row, column=2, value=f"¥{min(prices):,} - ¥{max(prices):,}")
        row += 1
        ws2.cell(row=row, column=1, value="平均价格")
        ws2.cell(row=row, column=2, value=f"¥{statistics.mean(prices):,.0f}")
        row += 1
        ws2.cell(row=row, column=1, value="中位数价格")
        ws2.cell(row=row, column=2, value=f"¥{statistics.median(prices):,.0f}")
        row += 1
        ws2.cell(row=row, column=1, value="高价商品(≥¥10,000)")
        ws2.cell(row=row, column=2, value=f"{high_price_count} 个 ({high_price_count/total*100:.1f}%)")
        
        row += 2
        ws2.cell(row=row, column=1, value="价格区间分布").font = Font(bold=True)
        row += 1
        ws2.cell(row=row, column=1, value="区间")
        ws2.cell(row=row, column=2, value="数量")
        ws2.cell(row=row, column=3, value="占比")
        row += 1
        for range_name, count in price_ranges.items():
            percentage = count / total * 100
            ws2.cell(row=row, column=1, value=range_name)
            ws2.cell(row=row, column=2, value=count)
            ws2.cell(row=row, column=3, value=f"{percentage:.1f}%")
            row += 1
        
        wb.save(excel_file)
        print(f"✓ 已生成Excel文件: {excel_file}")
        
    except ImportError:
        print("⚠ 未安装openpyxl库，无法生成Excel文件")
    
    print("\n" + "=" * 80)
    print("分析完成！")
    print("=" * 80)

if __name__ == '__main__':
    main()
