#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成螃蟹账号商品分析Excel报告
"""

import json
import re
from collections import Counter
import sys

# 检查openpyxl是否安装
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 需要安装openpyxl库")
    print("请运行: pip3 install openpyxl")
    sys.exit(1)

def load_data(file_path):
    """加载JSON数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def analyze_products(products):
    """分析商品数据"""
    # 价格统计
    prices = [p['price'] for p in products]
    
    price_ranges = {
        '0-500': 0,
        '500-1000': 0,
        '1000-5000': 0,
        '5000-10000': 0,
        '10000-50000': 0,
        '50000+': 0
    }
    
    for price in prices:
        if price <= 500:
            price_ranges['0-500'] += 1
        elif price <= 1000:
            price_ranges['500-1000'] += 1
        elif price <= 5000:
            price_ranges['1000-5000'] += 1
        elif price <= 10000:
            price_ranges['5000-10000'] += 1
        elif price <= 50000:
            price_ranges['10000-50000'] += 1
        else:
            price_ranges['50000+'] += 1
    
    # 平台统计
    platforms = [p['platform'] for p in products]
    platform_counts = Counter(platforms)
    
    # 命名特征
    names = [p['name'] for p in products]
    name_lengths = [len(name) for name in names]
    single_char_ids = [name for name in names if len(name) == 1]
    double_char_ids = [name for name in names if len(name) == 2 or '双字' in name]
    
    # 风格分类
    poetic_patterns = ['玄', '梵', '清', '月', '烟', '舞', '流', '脉', '广陵']
    poetic_names = [name for name in names if any(p in name for p in poetic_patterns)]
    
    cute_patterns = ['小', '萌', '甜', '妹', '宝宝', '仙女']
    cute_names = [name for name in names if any(p in name for p in cute_patterns)]
    
    aggressive_patterns = ['斩', '天', '帝皇', '神', '太子', '天下第一']
    aggressive_names = [name for name in names if any(p in name for p in aggressive_patterns)]
    
    celebrity_keywords = ['刘亦菲', '肖战', '林俊杰', '马斯克', '马龙']
    celebrity_names = [name for name in names if any(kw in name for kw in celebrity_keywords)]
    
    return {
        'price_stats': {
            'min': min(prices),
            'max': max(prices),
            'avg': sum(prices) / len(prices),
            'median': sorted(prices)[len(prices)//2],
            'high_value_count': len([p for p in prices if p >= 10000]),
            'ranges': price_ranges
        },
        'platform_stats': {
            'qq': platform_counts.get('QQ', 0),
            'wechat': platform_counts.get('微信', 0),
            'qq_pct': platform_counts.get('QQ', 0) / len(products) * 100,
            'wechat_pct': platform_counts.get('微信', 0) / len(products) * 100
        },
        'name_stats': {
            'avg_length': sum(name_lengths) / len(name_lengths),
            'single_char': len(single_char_ids),
            'double_char': len(double_char_ids),
            'poetic': len(poetic_names),
            'cute': len(cute_names),
            'aggressive': len(aggressive_names),
            'celebrity': len(celebrity_names)
        }
    }

def create_excel_report(products, analysis, output_path):
    """创建Excel报告"""
    wb = Workbook()
    
    # 定义样式
    header_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_font = Font(name='微软雅黑', bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: 概览
    ws1 = wb.active
    ws1.title = "分析概览"
    
    # 标题
    ws1.merge_cells('A1:B1')
    ws1['A1'] = "螃蟹账号《王者荣耀世界》商品分析报告"
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = center_align
    
    # 基本信息
    ws1['A3'] = "分析时间"
    ws1['B3'] = "2026-03-31"
    ws1['A4'] = "样本数量"
    ws1['B4'] = f"{len(products)}个商品"
    ws1['A5'] = "数据来源"
    ws1['B5'] = "pxb7.com"
    
    for row in range(3, 6):
        ws1[f'A{row}'].border = border
        ws1[f'B{row}'].border = border
        ws1[f'A{row}'].alignment = center_align
        ws1[f'B{row}'].alignment = center_align
    
    # 价格统计
    ws1['A7'] = "价格统计"
    ws1['A7'].font = header_font
    ws1['A7'].fill = header_fill
    ws1['A7'].border = border
    ws1['A7'].alignment = center_align
    
    price_data = [
        ("最低价格", f"¥{analysis['price_stats']['min']}"),
        ("最高价格", f"¥{analysis['price_stats']['max']:,}"),
        ("平均价格", f"¥{analysis['price_stats']['avg']:,.0f}"),
        ("中位数价格", f"¥{analysis['price_stats']['median']:,}"),
        ("高价商品(≥¥10,000)", f"{analysis['price_stats']['high_value_count']}个 ({analysis['price_stats']['high_value_count']/len(products)*100:.1f}%)")
    ]
    
    for i, (label, value) in enumerate(price_data, start=8):
        ws1[f'A{i}'] = label
        ws1[f'B{i}'] = value
        ws1[f'A{i}'].border = border
        ws1[f'B{i}'].border = border
        ws1[f'A{i}'].alignment = center_align
        ws1[f'B{i}'].alignment = center_align
    
    # 平台分布
    ws1['A14'] = "平台分布"
    ws1['A14'].font = header_font
    ws1['A14'].fill = header_fill
    ws1['A14'].border = border
    ws1['A14'].alignment = center_align
    
    ws1['A15'] = "QQ平台"
    ws1['B15'] = f"{analysis['platform_stats']['qq']}个 ({analysis['platform_stats']['qq_pct']:.1f}%)"
    ws1['A16'] = "微信平台"
    ws1['B16'] = f"{analysis['platform_stats']['wechat']}个 ({analysis['platform_stats']['wechat_pct']:.1f}%)"
    
    for row in range(15, 17):
        ws1[f'A{row}'].border = border
        ws1[f'B{row}'].border = border
        ws1[f'A{row}'].alignment = center_align
        ws1[f'B{row}'].alignment = center_align
    
    # 命名特征
    ws1['A18'] = "命名特征"
    ws1['A18'].font = header_font
    ws1['A18'].fill = header_fill
    ws1['A18'].border = border
    ws1['A18'].alignment = center_align
    
    name_data = [
        ("平均名称长度", f"{analysis['name_stats']['avg_length']:.1f}字符"),
        ("单字ID数量", f"{analysis['name_stats']['single_char']}个 ({analysis['name_stats']['single_char']/len(products)*100:.1f}%)"),
        ("双字ID数量", f"{analysis['name_stats']['double_char']}个 ({analysis['name_stats']['double_char']/len(products)*100:.1f}%)"),
        ("诗意/文学类", f"{analysis['name_stats']['poetic']}个"),
        ("可爱/萌系", f"{analysis['name_stats']['cute']}个"),
        ("霸气/中二类", f"{analysis['name_stats']['aggressive']}个"),
        ("名人/明星相关", f"{analysis['name_stats']['celebrity']}个")
    ]
    
    for i, (label, value) in enumerate(name_data, start=19):
        ws1[f'A{i}'] = label
        ws1[f'B{i}'] = value
        ws1[f'A{i}'].border = border
        ws1[f'B{i}'].border = border
        ws1[f'A{i}'].alignment = center_align
        ws1[f'B{i}'].alignment = center_align
    
    # 调整列宽
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 25
    
    # Sheet 2: 价格区间分布
    ws2 = wb.create_sheet("价格分布")
    
    ws2['A1'] = "价格区间分布"
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = center_align
    ws2.merge_cells('A1:C1')
    
    # 表头
    headers = ["价格区间", "数量", "占比"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 数据
    ranges = analysis['price_stats']['ranges']
    range_data = [
        ("0-500元", ranges['0-500'], f"{ranges['0-500']/len(products)*100:.1f}%"),
        ("500-1000元", ranges['500-1000'], f"{ranges['500-1000']/len(products)*100:.1f}%"),
        ("1000-5000元", ranges['1000-5000'], f"{ranges['1000-5000']/len(products)*100:.1f}%"),
        ("5000-10000元", ranges['5000-10000'], f"{ranges['5000-10000']/len(products)*100:.1f}%"),
        ("10000-50000元", ranges['10000-50000'], f"{ranges['10000-50000']/len(products)*100:.1f}%"),
        ("50000元以上", ranges['50000+'], f"{ranges['50000+']/len(products)*100:.1f}%")
    ]
    
    for row_idx, data in enumerate(range_data, start=4):
        for col_idx, value in enumerate(data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center_align
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 10
    
    # Sheet 3: 原始数据
    ws3 = wb.create_sheet("原始数据")
    
    ws3['A1'] = "商品原始数据"
    ws3['A1'].font = title_font
    ws3['A1'].fill = title_fill
    ws3['A1'].alignment = center_align
    ws3.merge_cells('A1:E1')
    
    # 表头
    headers = ["序号", "商品名称", "平台", "价格(元)", "发布时间"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 数据
    for idx, product in enumerate(products, start=4):
        ws3.cell(row=idx, column=1, value=product['id']).border = border
        ws3.cell(row=idx, column=2, value=product['name']).border = border
        ws3.cell(row=idx, column=3, value=product['platform']).border = border
        ws3.cell(row=idx, column=4, value=product['price']).border = border
        ws3.cell(row=idx, column=5, value=product['publish_time']).border = border
        
        for col in range(1, 6):
            ws3.cell(row=idx, column=col).alignment = center_align
    
    for col in range(1, 6):
        if col == 2:
            ws3.column_dimensions[chr(64+col)].width = 20
        else:
            ws3.column_dimensions[chr(64+col)].width = 12
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel报告已生成: {output_path}")

if __name__ == '__main__':
    products = load_data('/Users/jiewen/.real/users/user-777da3823a68e71779b041c8e7807df0/workspace/pxb7_products_100.json')
    analysis = analyze_products(products)
    create_excel_report(products, analysis, '/Users/jiewen/.real/users/user-777da3823a68e71779b041c8e7807df0/workspace/pxb7_products_analysis.xlsx')
